from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from scrapy import Selector
import time
import os
from urllib.parse import urljoin
from openpyxl import Workbook, load_workbook
import getpass

class linkedinApi:
    def __init__(self):
        self.username = input('\tEnter username: ')
        self.password = getpass.getpass(prompt='\tEnter Password: ', stream=None)
        self.done = []
        self.total_scraped = 0
        s = Service(ChromeDriverManager().install())
        self.driver = webdriver.Chrome(service=s)
        self.driver.maximize_window()
        self.filename = 'connections.xlsx'
        title = ['Profile Url','First Name', 'Last Name', 'Email', 'Phone', 'Website','Company', 'Position', 'Birthday','Address','Twitter','Connected']
        if os.path.exists(self.filename):
            print(f'\t{self.filename} already exists')
            self.wb = load_workbook(self.filename)
            self.ws = self.wb.active
            for x in self.ws['A'][1:]:
                self.done.append(x.value)
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.ws.append(title)

    def login(self):
        self.driver.get('https://www.linkedin.com/login')
        username = self.driver.find_element(By.ID, 'username')
        username.send_keys(self.username)

        password = self.driver.find_element(By.ID, 'password')
        password.send_keys(self.password)

        signIn = self.driver.find_element(By.CLASS_NAME, 'from__button--floating')
        signIn.click()

        time.sleep(2)
        try:
            response = Selector(text=self.driver.page_source)
            welcome = response.css('.block .t-bold::text').get().strip()
            print(f'\tWelcome {welcome}')
        except:
            time.sleep(60)
            self.login()
    
    def scroll(self, total):
        SCROLL_PAUSE_TIME = 5
        last_height = self.driver.execute_script("return document.body.scrollHeight")
        scroll_more = True
        while scroll_more:
            response = Selector(text=self.driver.page_source)
            conList = response.css('.artdeco-list')
            for con in conList:
                profileUrl = 'https://www.linkedin.com' + con.css('a::attr(href)').get()
                if profileUrl in self.done:
                    scroll_more = False
                    break
            print(f'\tScrolling down to load connections: {len(conList)}/{total}', end='\r')

            # Scroll down to bottom
            self.driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")

            # Wait to load page
            time.sleep(SCROLL_PAUSE_TIME)
            # Calculate new scroll height and compare with last scroll height
            new_height = self.driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                try:
                    load_more = WebDriverWait(self.driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, ".scaffold-finite-scroll__load-button"))
                    )
                    load_more.click()
                except:
                    break
            last_height = new_height
        print('/n')

    def contact_info(self,url):
        contact_url = urljoin(url, 'overlay/contact-info')
        self.driver.get(contact_url)
        WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'artdeco-modal__content'))
        )
        response = Selector(text=self.driver.page_source)
        address = response.css('.ci-address .pv-contact-info__ci-container a::attr(href)').get()
        try:
            email = response.css('.ci-email .pv-contact-info__ci-container a::attr(href)').get()
            email = email.split(':')[1]
        except:
            email = None
        twitter = response.css('.ci-twitter .pv-contact-info__ci-container a::attr(href)').get()
        try:
            birthday = response.css('.ci-birthday .pv-contact-info__ci-container span::text').get().strip()
        except:
            birthday = None
        try:
            connected = response.css('.ci-connected .pv-contact-info__ci-container span::text').get().strip() 
        except:
            connected = None
        website = response.css('.ci-websites .pv-contact-info__ci-container a::attr(href)').get()
        phone = response.css('.ci-phone .pv-contact-info__ci-container span::attr(href)').get()
        return address, email, twitter, birthday, connected, website, phone

    def profile_details(self,url):
        self.driver.get(url)
        time.sleep(2)
        response = Selector(text=self.driver.page_source)
        try:
            name = response.css('.pv-text-details__left-panel .text-heading-xlarge::text').get().strip()
        except:
            name = None
        try:
            position = response.css('.pv-text-details__left-panel .text-body-medium::text').get().strip()
        except:
            position = None
        try:
            company = response.css('.pv-text-details__right-panel-item .inline-show-more-text::text').get().strip()
        except:
            company = None
        address, email, twitter, birthday, connected, phone, website = self.contact_info(url)
        item = {
            'Profile Url': url,
            'First Name': name.split()[0],
            'Last Name': ' '.join(name.split()[1:]),
            'Email': email,
            'Phone': phone,
            'Website': website,
            'Company': company,
            'Position': position,
            'Birthday': birthday,
            'Address': address,
            'Twitter': twitter,
            'Connected': connected,
        }
        self.total_scraped += 1
        print(f'\t{self.total_scraped}: {url}')
        self.process_item(item)

    def process_item(self, item):
        line = [item['Profile Url'], item['First Name'], item['Last Name'], item['Email'], item['Phone'], item['Website'], item['Company'], item['Position'], item['Birthday'], item['Address'], item['Twitter'],item['Connected']]
        self.ws.append(line)
        self.wb.save(self.filename)

    def get_connection(self):
        self.driver.get('https://www.linkedin.com/mynetwork/invite-connect/connections/')
        time.sleep(0.5)
        total_connections = WebDriverWait(self.driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "#ember34 .t-black.t-normal"))
        ).text
        total_connections = total_connections.split()[0]
        if True:
            self.scroll(total_connections)
        response = Selector(text=self.driver.page_source)
        connectionDivs = response.css('.artdeco-list')
        for connectionDiv in connectionDivs:
            profileUrl = 'https://www.linkedin.com' + connectionDiv.css('a::attr(href)').get()
            if profileUrl in self.done:
                continue
            self.profile_details(profileUrl)

    def main(self):
        self.login()
        self.get_connection()
        self.driver.quit()
        if self.total_scraped == 0:
            print(f'\tNo new connections added.')
        print(f'\tOutput is saved as {self.filename}')

if __name__=='__main__':
    Lapi = linkedinApi()
    Lapi.main()